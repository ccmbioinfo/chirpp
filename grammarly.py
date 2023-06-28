import pandas as pd
from gingerit.gingerit import GingerIt
import re
from openpyxl import load_workbook

def text_fix(fileName, sheetName, textCol):
    """
        Corrects and saves the text in the specified column of an Excel sheet using GingerIt and regex.

        Args:
            fileName (str): Path to the Excel file.
            sheetName (str): Name of the sheet containing the text column.
            textCol (str): Name of the column containing the text to be checked.

        Returns:
            None
    """
    
    #Text parser
    parser = GingerIt()

    # Path to the Excel file
    excel_file = fileName

    # Name of the sheet containing the text column
    sheet_name = sheetName

    # Name of the column containing the text to be checked
    text_column = textCol

    # Read the Excel file into a DataFrame
    df = pd.read_excel(excel_file, sheet_name=sheet_name)

    book = load_workbook(fileName)
    sheet2 = book[sheetName]

    for index, row in df.iterrows():
        test = row[text_column]
        if isinstance(test, str):
            test = str(row[text_column]).lower()  # Convert to string to handle NaN values

            # Split the text into blocks of 450 characters
            blocks = [test[i:i+450] for i in range(0, len(test), 450)]

            # Apply the grammar checker to each block of text
            block_results = []
            for block in blocks:
                result = parser.parse(block)
                block_results.append(result['result'])

            # All regex stuff below
            # Concatenate the block results
            combined_result = "".join(block_results)

            # Remove extra periods and random new lines
            corrected = re.sub(r'\.{2,}', '.', combined_result)
            corrected = re.sub(r'\n+', ' ', corrected)

            # Fix " Presented to Ed immediately arrived by 16 : 30"
            corrected = re.sub(r'(?<=\d)\s*:\s*(?=\d)', ':', corrected)

            # Remove space behind the first "
            corrected = re.sub(r'"\s+', r'"', corrected, count=1)

            # Remove space in front of the last "
            corrected = re.sub(r'\s+"$', r'"', corrected)

            # Remove "-" without adding a space
            corrected = re.sub(r'[-–—](?=\s)', '', corrected)

            # Remove ">" symbol
            corrected = corrected.replace(">", "").replace("<", "")

            # Remove spaces between words and punctuation that aren't periods or commas
            corrected = re.sub(r'(?<=[^\s.,])\s{2,}(?=[^\s.,])', ' ', corrected)

            # Remove sentences that are only 1 word long
            corrected = ' '.join(sentence for sentence in corrected.split('.') if len(sentence.split()) > 1)

            # Add 1 space behind periods that don't have one
            corrected = re.sub(r'(?<=[^.])\.(?=\S)', '. ', corrected)

            # If there is more than 1 space after a word, add a period after that word and keep 1 space before the next word
            corrected = re.sub(r'(?<=\w)\s{2,}(?=\w)', '. ', corrected) 
 
            result = "".join(corrected)
            column_index = df.columns.get_loc(textCol) + 1
            sheet2.cell(row=index+2, column=column_index).value = result

    # Save the changes to the workbook
    book.save(fileName)

textFix("C:/Users/Kevin Yao/Desktop/grammar_small.xlsx", "Sheet1", "testNars")
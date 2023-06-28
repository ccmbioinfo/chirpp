import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import NamedStyle

def moveOver(fileName, sheetFrom, sheetTo):
    """
    Moves rows from one sheet to another in an Excel workbook based on specific criteria.

    Args:
        fileName (str): The path of the Excel file.
        sheetFrom (str): The name of the sheet from which rows will be moved.
        sheetTo (str): The name of the sheet to which rows will be moved.

    Returns:
        None

    Raises:
        None

    Description:
        The function reads an Excel file specified by 'fileName' and moves rows from the 'sheetFrom' to 'sheetTo'
        based on specific criteria. The rows are moved if the 'Chief Complaint' column contains keywords from the
        'Orange' column of the 'Sheet1' in the file 'Chief_Complaint.xlsx'. The rows that meet the criteria are
        highlighted with a yellow color and deleted from 'sheetFrom'.

        The function utilizes the openpyxl library to manipulate Excel files.

    Example:
        moveOver('data.xlsx', 'Sheet1', 'Sheet2')
    """
    filename = fileName

    df = pd.read_excel('C:/Users/Kevin Yao/Desktop/Yellow.xlsx', sheet_name='Sheet1')
    df0 = pd.read_excel('C:/Users/Kevin Yao/Desktop/Red.xlsx', sheet_name='Sheet1')
    complaint_listY = df['Yellow'].tolist()
    complaint_listR = df0['Red'].tolist()

    complaint_listYe = [keyword.lower().replace(' ', '') for keyword in complaint_listY]
    complaint_listRe = [keyword.lower().replace(' ', '') for keyword in complaint_listR]

    df1 = pd.read_excel(filename, sheet_name=sheetFrom)
    df2 = pd.read_excel(filename, sheet_name=sheetTo)

    book = load_workbook(filename)
    sheet1 = book[sheetFrom]
    sheet2 = book[sheetTo]

    counter = 0
    df1['Chief Complaint'] = df1['Chief Complaint'].str.lower().str.replace(' ', '')
    # Create a named style for highlighting the entire row
    highlight_row = NamedStyle(name="highlight_row")
    highlight_row.fill = PatternFill(fill_type="solid", fgColor="FFFF00")

    delete_rows = []

    counter = 0 
    # Highlight and move rows from Sheet1 to Sheet2
    for index, row in df1.iterrows():
        if isinstance(row['Chief Complaint'], str):
            cc = row['Chief Complaint'].lower()
            if any(keyword in cc for keyword in complaint_listRe):
                delete_rows.append(index)
                counter += 1
                print(counter)
                sheet2.append(row.values.tolist())
                row_number = sheet2.max_row
                for cell in sheet2[row_number]:
                    cell.style = highlight_row

            # elif any(keyword in cc for keyword in complaint_listYe):
            #     skN = str(row['SK Narrative']).lower()  # Convert to string to handle NaN values
            #     if "fall" in skN or "fell" in skN or ("trauma" in skN and "atraumatic" not in skN and "no trauma" not in skN and "no known trauma" not in skN) or ("injury" in skN and "no injury" not in skN and "no known injur" not in skN) or " hit " in skN or "bump" in skN or "slip" in skN or " digging" in skN or "poking" in skN or "cut" in skN:
            #         delete_rows.append(index)
            #         counter += 1
            #         print(counter)
            #         sheet2.append(row.values.tolist())
            #         row_number = sheet2.max_row
            #         for cell in sheet2[row_number]:
            #             cell.style = highlight_row


    # Delete rows from Sheet1
    for index in delete_rows[::-1]:
        sheet1.delete_rows(index + 2)

    # Save the changes to the workbook
    book.save(filename)

moveOver("C:/Users/Kevin Yao/Desktop/Aug 2022.xlsx", "Sheet1", "Sheet2")
